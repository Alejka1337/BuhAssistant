import csv
import chardet
import re
from io import StringIO, BytesIO
from typing import List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import xlrd
from app.schemas.tax_requisite import TaxRequisiteCreate
from app.models.tax_requisite import TaxRequisiteType


def detect_encoding(file_content: bytes) -> str:
    """
    Визначити кодування CSV файлу
    Спробувати кілька варіантів кодування
    """
    # Список кодувань для спроби (від найбільш ймовірних)
    encodings_to_try = ['utf-8', 'windows-1251', 'cp1251', 'utf-16', 'latin1']
    
    # Спробувати chardet спочатку
    result = chardet.detect(file_content)
    if result['encoding'] and result['confidence'] > 0.7:
        encodings_to_try.insert(0, result['encoding'])
    
    # Спробувати кожне кодування
    for encoding in encodings_to_try:
        try:
            file_content.decode(encoding)
            return encoding
        except (UnicodeDecodeError, AttributeError):
            continue
    
    # Якщо нічого не вийшло, використовуємо latin1 (завжди працює, але може бути неправильно)
    return 'latin1'


def detect_delimiter(lines: List[str]) -> str:
    """
    Визначити роздільник CSV (кома, табуляція або крапка з комою)
    Перевіряється перша непорожня лінія
    """
    for line in lines:
        if line.strip():
            if '\t' in line:
                return '\t'
            elif ';' in line:
                return ';'
            else:
                return ','
    return ','


def extract_district_from_recipient(recipient_name: str) -> Tuple[str, Optional[str]]:
    """
    Витягнути назву району/міста/села з назви отримувача
    
    Приклад:
    - "ГУК у Він.обл./с.Агрономiчне/11010100" -> ("ГУК у Він.обл.", "с.Агрономiчне")
    - "ГУК у Він.обл./м.Бар/21081500" -> ("ГУК у Він.обл.", "м.Бар")
    - "ГУК у Він.обл./смт Сутиски/11010100" -> ("ГУК у Він.обл.", "смт Сутиски")
    - "Донецьке ГУК/Авдiївська МТГ/11010100" -> ("Донецьке ГУК", "Авдiївська МТГ")
    - "Донецьке ГУК/Андрiївська СТГ/11010400" -> ("Донецьке ГУК", "Андрiївська СТГ")
    - "ГУК у Він.обл./Він.обл/11011000" -> ("ГУК у Він.обл.", None)  # область, не місто
    - "ГУ ДПС у Вінницькій обл." -> ("ГУ ДПС у Вінницькій обл.", None)
    """
    # Шукаємо паттерн: щось / місто, село, смт або територіальна громада / код
    # Регулярний вираз для:
    # 1. м.Назва, с.Назва або смт Назва
    # 2. Назва МТГ/СТГ/ОТГ (міська/сільська/об'єднана територіальна громада)
    # НЕ для "Він.обл" чи інших назв областей
    
    # Спробувати знайти формат з префіксом (м., с., смт)
    match = re.search(r'/\s*(м\.|с\.|смт\s+)([^/]+?)(?:/|$)', recipient_name)
    if match:
        district = f"{match.group(1)}{match.group(2)}".strip()
        base_name = recipient_name.split('/')[0].strip()
        return (base_name, district)
    
    # Спробувати знайти формат з суфіксом громади (МТГ, СТГ, ОТГ)
    match = re.search(r'/\s*([А-ЯЁІЇЄҐа-яёіїєґ\'\-і]+\s+(?:МТГ|СТГ|ОТГ|мтг|стг|отг))(?:/|$)', recipient_name, re.IGNORECASE)
    if match:
        district = match.group(1).strip()
        base_name = recipient_name.split('/')[0].strip()
        return (base_name, district)
    
    # Якщо не знайшли жодного патерну - це область або немає district
    # Повертаємо як є (district = None)
    return (recipient_name.split('/')[0].strip() if '/' in recipient_name else recipient_name, None)


def parse_esv_csv(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Парсинг CSV файлу з реквізитами ЄСВ
    
    Очікувані колонки:
    - Назва банку / Банк отримувача
    - Назва органу ДПС
    - Код за ЄДРПОУ органу ДПС
    - Номер рахунку (IBAN)
    - Символ звітності
    - Категорії платників єдиного внеску
    """
    encoding = detect_encoding(file_content)
    content = file_content.decode(encoding)
    
    # Видалити BOM якщо є
    if content.startswith('\ufeff'):
        content = content[1:]
    
    # Розбити на рядки
    lines = content.strip().split('\n')
    
    # Знайти рядок з заголовками (містить "Символ звітності")
    header_line_idx = None
    for idx, line in enumerate(lines):
        if 'Символ звітності' in line or 'символ звітності' in line.lower():
            header_line_idx = idx
            break
    
    if header_line_idx is None:
        raise ValueError("Не знайдено рядок з заголовками (повинен містити 'Символ звітності')")
    
    # Створити CSV reader починаючи з рядка заголовків
    csv_content = '\n'.join(lines[header_line_idx:])
    csv_file = StringIO(csv_content)
    
    # Автоматичне визначення роздільника
    delimiter = detect_delimiter(lines)
    
    reader = csv.DictReader(csv_file, delimiter=delimiter)
    
    requisites = []
    
    for row in reader:
        # Пропустити порожні рядки
        if not any(row.values()):
            continue
            
        # Знайти колонку з символом звітності (гнучкий пошук)
        symbol = ''
        for key in row.keys():
            if 'символ звітності' in key.lower():
                symbol = row[key].strip()
                break
        
        # Фільтр: тільки 201 та 204
        if symbol not in ['201', '204']:
            continue
        
        # Визначити тип (використовуємо .value для отримання строкового значення)
        requisite_type = TaxRequisiteType.ESV_FOP.value if symbol == '201' else TaxRequisiteType.ESV_EMPLOYEES.value
        
        # Знайти інші колонки (гнучкий пошук)
        bank_name = ''
        recipient_name = ''
        recipient_code = ''
        iban = ''
        description = ''
        
        for key, value in row.items():
            key_lower = key.lower()
            if 'банк' in key_lower and 'назва' in key_lower:
                bank_name = value.strip()
            elif 'назва органу дпс' in key_lower or 'назва органу' in key_lower:
                recipient_name = value.strip()
            elif 'код за єдрпоу' in key_lower:
                recipient_code = value.strip()
            elif 'номер рахунку' in key_lower or 'iban' in key_lower:
                iban = value.strip()
            elif 'категорії платників' in key_lower:
                description = value.strip()
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=None,  # В малих файлах немає району
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=symbol,
            description=description
        )
        
        requisites.append(requisite)
    
    return requisites


def parse_tax_csv(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Парсинг CSV файлу з реквізитами для інших податків
    
    Очікувані колонки:
    - Код обл.
    - Найменування адміністративно-териториальної одиниці України
    - Отримувач (найменування органу Казначейства)
    - Код отримувача (ЄДРПОУ)
    - Банк отримувача
    - Номер рахунку (IBAN)
    - Код класифікації доходів бюджету
    - Найменування коду класифікації доходів бюджету
    """
    encoding = detect_encoding(file_content)
    content = file_content.decode(encoding)
    
    # Видалити BOM якщо є
    if content.startswith('\ufeff'):
        content = content[1:]
    
    # Автоматичне визначення роздільника
    lines = content.split('\n')
    delimiter = detect_delimiter(lines)
    
    # Знайти перший рядок даних (починається з коду області, наприклад "02")
    # Заголовок буде рядок перед першими даними
    header_line_idx = 0
    data_start_idx = 0
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue
        # Шукаємо рядок який починається з цифр (код області)
        parts = stripped.split(delimiter)
        if parts and parts[0].strip().isdigit():
            data_start_idx = i
            # Заголовок - попередній непорожній рядок
            for j in range(i - 1, -1, -1):
                if lines[j].strip():
                    header_line_idx = j
                    break
            break
    
    # Створити CSV з заголовка і даних
    if data_start_idx > 0:
        csv_content = '\n'.join([lines[header_line_idx]] + lines[data_start_idx:])
    else:
        # Якщо не знайшли дані, використовуємо весь файл
        csv_content = content
    
    csv_file = StringIO(csv_content)
    reader = csv.DictReader(csv_file, delimiter=delimiter)
    
    requisites = []
    
    # Мапінг кодів класифікації до типів (використовуємо .value для строкових значень)
    code_to_type = {
        '11010100': TaxRequisiteType.PDFO_EMPLOYEES.value,
        '11011000': TaxRequisiteType.MILITARY_EMPLOYEES.value,
        '11011700': TaxRequisiteType.MILITARY_FOP.value,
        '18050400': TaxRequisiteType.SINGLE_TAX_FOP.value,
    }
    
    # Відладочний вивід заголовків (перший рядок)
    first_row = True
    
    for row in reader:
        # Показати заголовки з першого рядка
        if first_row:
            print(f"🔍 CSV Headers: {list(row.keys())}")
            first_row = False
        
        # Пропустити порожні рядки
        if not any(row.values()):
            continue
        
        # Отримати код класифікації (гнучкий пошук назви колонки)
        classification_code = ''
        for key in row.keys():
            if 'Код класифікації' in key or 'код класифікації' in key:
                classification_code = row[key].strip()
                break
        
        # Фільтр: тільки потрібні коди
        if classification_code not in code_to_type:
            continue
        
        # Визначити тип
        requisite_type = code_to_type[classification_code]
        
        # Гнучкий пошук для всіх колонок
        recipient_name = ''
        recipient_code = ''
        bank_name = ''
        iban = ''
        description = ''
        district_code = ''
        district_name = ''
        
        for key, value in row.items():
            key_lower = key.lower()
            if 'отримувач' in key_lower and 'код' not in key_lower:
                recipient_name = value.strip()
            elif 'код отримувача' in key_lower or 'єдрпоу' in key_lower:
                recipient_code = value.strip()
            elif 'банк' in key_lower and 'отримувача' in key_lower:
                bank_name = value.strip()
            elif 'номер рахунку' in key_lower or 'iban' in key_lower:
                iban = value.strip()
            elif 'найменування коду' in key_lower:
                description = value.strip()
            elif 'код' in key_lower and 'обл' in key_lower:
                district_code = value.strip()
            elif 'назва' in key_lower and ('район' in key_lower or 'громад' in key_lower):
                district_name = value.strip()
        
        # Витягнути district з recipient_name (наприклад: "ГУК у Він.обл./м.Бар/...")
        base_recipient_name, extracted_district = extract_district_from_recipient(recipient_name)
        
        # Використовуємо витягнутий district або комбінацію з колонок
        if extracted_district:
            district = extracted_district
            recipient_name = base_recipient_name
        else:
            district = f"{district_code} {district_name}".strip() if district_code or district_name else None
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=district,
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=classification_code,
            description=description
        )
        
        requisites.append(requisite)
    
    return requisites


def parse_esv_xlsx(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Парсинг XLSX файлу з реквізитами ЄСВ
    
    Очікувані колонки:
    - Назва банку / Банк отримувача
    - Назва органу ДПС
    - Код за ЄДРПОУ органу ДПС
    - Номер рахунку (IBAN)
    - Символ звітності
    - Категорії платників єдиного внеску
    """
    # Завантажити Excel файл
    workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    
    # Знайти рядок з заголовками (містить "Символ звітності")
    header_row_idx = None
    headers = []
    
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        # Перевірити чи містить рядок "Символ звітності"
        for cell in row:
            if cell and isinstance(cell, str) and 'символ звітності' in cell.lower():
                header_row_idx = idx
                headers = [str(cell).strip() if cell else '' for cell in row]
                break
        if header_row_idx:
            break
    
    if not header_row_idx:
        raise ValueError("Не знайдено рядок з заголовками (повинен містити 'Символ звітності')")
    
    print(f"🔍 XLSX Headers (ESV): {headers}")
    
    requisites = []
    
    # Обробити всі рядки після заголовка
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Пропустити порожні рядки
        if not any(row):
            continue
        
        # Створити словник з даними
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = str(value).strip() if value else ''
        
        # Знайти символ звітності
        symbol = ''
        for key, value in row_data.items():
            if 'символ звітності' in key.lower():
                symbol = value
                break
        
        # Фільтр: тільки 201 та 204
        if symbol not in ['201', '204']:
            continue
        
        # Визначити тип
        requisite_type = TaxRequisiteType.ESV_FOP.value if symbol == '201' else TaxRequisiteType.ESV_EMPLOYEES.value
        
        # Знайти інші колонки
        bank_name = ''
        recipient_name = ''
        recipient_code = ''
        iban = ''
        description = ''
        
        for key, value in row_data.items():
            key_lower = key.lower()
            if 'банк' in key_lower and ('назва' in key_lower or 'отримувача' in key_lower):
                bank_name = value
            elif 'назва органу дпс' in key_lower or 'назва органу' in key_lower:
                recipient_name = value
            elif 'код за єдрпоу' in key_lower or ('код' in key_lower and 'єдрпоу' in key_lower):
                recipient_code = value
            elif 'номер рахунку' in key_lower or 'iban' in key_lower:
                iban = value
            elif 'категорії платників' in key_lower:
                description = value
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=None,
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=symbol,
            description=description
        )
        
        requisites.append(requisite)
    
    workbook.close()
    return requisites


def parse_tax_xlsx_simple(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Спрощений парсинг XLSX файлу з фіксованою структурою
    
    Структура файлу:
    - Рядок 1: Заголовки (A-G)
    - Рядки 2+: Дані
    
    Колонки:
    - A: Назва АТО (district)
    - B: Отримувач (ГУК у ... /район/код) - витягуємо тільки "ГУК у ..."
    - C: Код ЄДРПОУ (recipient_code)
    - D: Банк отримувача (bank_name)
    - E: Номер рахунку IBAN (iban)
    - F: Код класифікації (classification_code) - фільтр по потрібним
    - G: Найменування коду (description)
    """
    workbook = load_workbook(BytesIO(file_content), data_only=True)
    sheet = workbook.active
    
    requisites = []
    
    # Мапінг кодів класифікації до типів
    code_to_type = {
        '11010100': TaxRequisiteType.PDFO_EMPLOYEES.value,
        '11011000': TaxRequisiteType.MILITARY_EMPLOYEES.value,
        '11011700': TaxRequisiteType.MILITARY_FOP.value,
        '18050400': TaxRequisiteType.SINGLE_TAX_FOP.value,
    }
    
    # Мапінг скорочених назв до повних назв ГУК
    region_to_guk = {
        'Вінницька область': 'ГУК у Вінницькій області',
        'Волинська область': 'ГУК у Волинській області',
        'Дніпропетровська область': 'ГУК у Дніпропетровській області',
        'Донецька область': 'Донецьке ГУК',
        'Житомирська область': 'ГУК у Житомирській області',
        'Закарпатська область': 'ГУК у Закарпатській області',
        'Запорізька область': 'ГУК у Запорізькій області',
        'Івано-Франківська область': 'ГУК в Івано-Франківській області',
        'м. Київ': 'ГУК у м.Києві',
        'Київська область': 'ГУК у Київській області',
        'Кіровоградська область': 'ГУК у Кіровоградській області',
        'Луганська область': 'ГУК у Луганській області',
        'Львівська область': 'ГУК у Львівській області',
        'Миколаївська область': 'ГУК у Миколаївській області',
        'Одеська область': 'ГУК в Одеській області',
        'Полтавська область': 'ГУК у Полтавській області',
        'Рівненська область': 'ГУК у Рівненській області',
        'Сумська область': 'ГУК у Сумській області',
        'Тернопільська область': 'ГУК у Тернопільській області',
        'Харківська область': 'ГУК у Харківській області',
        'Херсонська область': 'ГУК у Херсонській області',
        'Хмельницька область': 'ГУК у Хмельницькій області',
        'Черкаська область': 'ГУК у Черкаській області',
        'Чернівецька область': 'ГУК у Чернівецькій області',
        'Чернігівська область': 'ГУК у Чернігівській області',
    }
    
    # Обробити всі рядки, починаючи з 2-го (індекс 2 в openpyxl, бо нумерація з 1)
    for row_idx in range(2, sheet.max_row + 1):
        row = sheet[row_idx]
        
        # Витягнути дані з фіксованих позицій
        district = str(row[0].value).strip() if row[0].value else ''
        recipient_name_full = str(row[1].value).strip() if row[1].value else ''
        recipient_code = str(row[2].value).strip().replace('.0', '') if row[2].value else ''
        bank_name = str(row[3].value).strip() if row[3].value else 'Казначейство України (ел. адм. подат.)'
        iban = str(row[4].value).strip() if row[4].value else ''
        classification_code = str(row[5].value).strip().replace('.0', '') if row[5].value else ''
        description = str(row[6].value).strip() if row[6].value else ''
        
        # Пропустити порожні рядки
        if not (district or recipient_name_full or iban):
            continue
        
        # Фільтр: тільки потрібні коди
        if classification_code not in code_to_type:
            continue
        
        # Визначити тип
        requisite_type = code_to_type[classification_code]
        
        # Витягнути recipient_name з повної назви (частина до першого /)
        # Наприклад: "ГУК у Дн-кiй обл/Iнгул.р/11010500" -> "ГУК у Дн-кiй обл"
        if '/' in recipient_name_full:
            recipient_name = recipient_name_full.split('/')[0].strip()
        else:
            recipient_name = recipient_name_full
        
        # Нормалізувати recipient_name до повної назви ГУК
        # Використовуємо мапінг region -> ГУК
        recipient_name = region_to_guk.get(region, recipient_name)
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=district,
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=classification_code,
            description=description
        )
        
        requisites.append(requisite)
    
    workbook.close()
    
    print(f"📊 Статистика парсингу (Simple XLSX):")
    print(f"  Фактично створено записів: {len(requisites)}")
    
    return requisites


def parse_tax_xlsx_OLD(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Парсинг XLSX файлу з реквізитами для інших податків
    
    Очікувані колонки:
    - Код обл.
    - Назва району/територіальної громади
    - Отримувач
    - Код отримувача (ЄДРПОУ)
    - Банк отримувача
    - Номер рахунку (IBAN)
    - Код класифікації доходів бюджету
    - Найменування коду класифікації доходів бюджету
    """
    # Завантажити Excel файл
    workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    
    # Знайти рядок з заголовками (перший рядок який містить очікувані назви колонок)
    # або перший рядок даних (починається з цифр - код області)
    header_row_idx = None
    headers = []
    
    # Спочатку шукаємо перший рядок даних (починається з коду області)
    first_data_row_idx = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if not any(row):
            continue
        
        first_cell = str(row[0]).strip() if row[0] else ''
        
        # Пропустити рядок нумерації (1.0, 2.0, 3.0, ...)
        if first_cell in ['1', '1.0', '№', 'N']:
            continue
        
        # Перевіряємо перший стовпець - чи це код області
        # Може бути: 5, 5.0, 02, 14, тощо
        # Очищаємо від .0 якщо є
        cleaned_cell = first_cell.replace('.0', '')
        
        # Код області - цифри (1-4 символи)
        if cleaned_cell.isdigit() and 1 <= len(cleaned_cell) <= 4:
            first_data_row_idx = idx
            print(f"🔍 Знайдено перший рядок даних на позиції {idx}, перша комірка: '{first_cell}'")
            
            # Шукаємо заголовок - попередній непорожній рядок, але НЕ рядок нумерації
            for prev_idx in range(idx - 1, 0, -1):
                prev_row = list(sheet.iter_rows(min_row=prev_idx, max_row=prev_idx, values_only=True))[0]
                if not any(prev_row):
                    continue
                    
                # Перевірити чи це не рядок нумерації
                prev_first_cell = str(prev_row[0]).strip() if prev_row[0] else ''
                if prev_first_cell in ['1', '1.0', '№', 'N']:
                    continue
                
                # Це заголовок!
                header_row_idx = prev_idx
                headers = [str(cell).strip() if cell else '' for cell in prev_row]
                break
            break
    
    if not header_row_idx or not headers:
        # Додатковий лог для діагностики
        print("🔍 Не вдалося знайти заголовки. Перші 10 рядків файлу:")
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            print(f"  Рядок {idx}: {[str(cell)[:50] if cell else '' for cell in row[:5]]}")
        raise ValueError("Не вдалося знайти заголовки в XLSX файлі")
    
    print(f"🔍 XLSX Headers (Tax): {headers}")
    
    requisites = []
    
    # Мапінг кодів класифікації до типів
    code_to_type = {
        '11010100': TaxRequisiteType.PDFO_EMPLOYEES.value,
        '11011000': TaxRequisiteType.MILITARY_EMPLOYEES.value,
        '11011700': TaxRequisiteType.MILITARY_FOP.value,
        '18050400': TaxRequisiteType.SINGLE_TAX_FOP.value,
    }
    
    # Лічильники для діагностики
    all_codes_found = []
    matched_codes_count = 0
    
    # Обробити всі рядки після заголовка
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Пропустити порожні рядки
        if not any(row):
            continue
        
        # Створити словник з даними
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = str(value).strip() if value else ''
        
        # Debug: показати перші рядки заголовків
        if len(all_codes_found) == 0:
            print(f"🔍 DEBUG Заголовки: {list(row_data.keys())}")
        
        # Отримати код класифікації
        classification_code = ''
        
        # Спробувати знайти в окремій колонці
        # Використовуємо startswith для точнішого пошуку
        for key, value in row_data.items():
            key_lower = key.lower().strip()
            # Шукаємо колонку що починається з "код класифікації" або містить тільки "код класифікації доходів"
            if key_lower.startswith('код класифікації'):
                classification_code = value
                # Видалити .0 якщо є (Excel зберігає числа як float)
                classification_code = str(classification_code).replace('.0', '').strip()
                if len(all_codes_found) < 3:
                    print(f"✅ DEBUG Знайдено код у колонці '{key}': '{classification_code}'")
                break
        
        # Якщо не знайшли в окремій колонці, спробувати витягнути з "Назва отримувача"
        # Формат: "Донецьке ГУК/Авдiївська МТГ/11010100"
        if not classification_code:
            recipient_name_temp = ''
            for key, value in row_data.items():
                key_lower = key.lower()
                if 'отримувач' in key_lower and 'код' not in key_lower and 'банк' not in key_lower:
                    recipient_name_temp = value
                    break
            
            # Витягнути код з кінця (після останнього /)
            if '/' in recipient_name_temp:
                parts = recipient_name_temp.split('/')
                last_part = parts[-1].strip()
                # Перевірити чи це код (8 цифр)
                if last_part.isdigit() and len(last_part) == 8:
                    classification_code = last_part
        
        # Зібрати всі знайдені коди для статистики
        if classification_code and len(all_codes_found) < 50:
            all_codes_found.append(classification_code)
        
        # Фільтр: тільки потрібні коди
        if classification_code not in code_to_type:
            continue
        
        matched_codes_count += 1
        
        # Визначити тип
        requisite_type = code_to_type[classification_code]
        
        # Гнучкий пошук для всіх колонок
        recipient_name = ''
        recipient_code = ''
        bank_name = ''
        iban = ''
        description = ''
        district_code = ''
        district_name = ''
        
        for key, value in row_data.items():
            key_lower = key.lower().strip()
            # Отримувач - колонка що містить "отримувач", але НЕ "код отримувача" і НЕ "банк отримувача"
            if 'отримувач' in key_lower and 'код' not in key_lower.split()[:2] and 'банк' not in key_lower.split()[:2]:
                recipient_name = value
            elif 'код отримувача' in key_lower or ('єдрпоу' in key_lower and 'код' in key_lower):
                recipient_code = value
            elif 'банк отримувача' in key_lower or 'банк' in key_lower.split()[:2]:
                bank_name = value
            elif 'номер рахунку' in key_lower or 'iban' in key_lower:
                iban = value
            elif 'найменування коду класифікації' in key_lower:
                description = value
            elif 'код' in key_lower and 'обл' in key_lower:
                district_code = value
            elif 'назва' in key_lower and ('район' in key_lower or 'громад' in key_lower or 'населеного' in key_lower):
                district_name = value
        
        # Витягнути district з recipient_name (наприклад: "ГУК у Він.обл./м.Бар/...")
        base_recipient_name, extracted_district = extract_district_from_recipient(recipient_name)
        
        # Використовуємо витягнутий district або комбінацію з колонок
        if extracted_district:
            district = extracted_district
            recipient_name = base_recipient_name
        else:
            district = f"{district_code} {district_name}".strip() if district_code or district_name else None
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            if len(requisites) < 3:
                print(f"⚠️  DEBUG Пропущено рядок: recipient_name='{recipient_name[:30] if recipient_name else ''}', recipient_code='{recipient_code}', iban='{iban}'")
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=district,
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=classification_code,
            description=description
        )
        
        requisites.append(requisite)
    
    # Вивести статистику
    print(f"📊 Статистика парсингу (XLSX):")
    print(f"  Всього рядків оброблено: {len(all_codes_found)}")
    print(f"  Знайдено потрібних кодів: {matched_codes_count}")
    print(f"  Фактично створено записів: {len(requisites)}")
    print(f"  Перші 50 кодів у файлі: {all_codes_found[:50]}")
    print(f"  Потрібні коди: {list(code_to_type.keys())}")
    
    workbook.close()
    return requisites


def is_xls_file(file_content: bytes) -> bool:
    """
    Визначити чи файл є старим форматом .xls
    Перевіряємо магічні байти для Composite Document File
    """
    # Composite Document File V2 починається з D0 CF 11 E0
    if len(file_content) < 8:
        return False
    return file_content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'


def parse_esv_xls(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Парсинг XLS файлу з реквізитами для ЄСВ (старий формат Excel)
    """
    workbook = xlrd.open_workbook(file_contents=file_content)
    sheet = workbook.sheet_by_index(0)
    
    # Знайти заголовки
    header_row_idx = None
    for row_idx in range(min(20, sheet.nrows)):
        row_values = [str(cell.value).strip() for cell in sheet.row(row_idx)]
        row_text = ' '.join(row_values).lower()
        
        if 'символ звітності' in row_text or 'символ' in row_text:
            header_row_idx = row_idx
            break
    
    if header_row_idx is None:
        raise ValueError("Не вдалося знайти заголовки в XLS файлі")
    
    # Отримати заголовки
    headers = [str(cell.value).strip() for cell in sheet.row(header_row_idx)]
    print(f"🔍 XLS Headers (ESV): {headers}")
    
    requisites = []
    
    # Обробити всі рядки після заголовка
    for row_idx in range(header_row_idx + 1, sheet.nrows):
        row = sheet.row(row_idx)
        row_data = {}
        
        for idx, cell in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = str(cell.value).strip()
        
        # Пропустити порожні рядки
        if not any(row_data.values()):
            continue
        
        # Знайти символ звітності
        symbol = ''
        for key, value in row_data.items():
            if 'символ' in key.lower():
                symbol = value.strip()
                break
        
        # Фільтр: тільки 201 та 204
        if symbol not in ['201', '204', '201.0', '204.0']:
            continue
        
        # Очистити символ від .0
        symbol = symbol.replace('.0', '')
        
        # Визначити тип
        if symbol == '201':
            requisite_type = TaxRequisiteType.ESV_FOP.value
        else:  # 204
            requisite_type = TaxRequisiteType.ESV_EMPLOYEES.value
        
        # Гнучкий пошук для всіх колонок
        recipient_name = ''
        recipient_code = ''
        bank_name = ''
        iban = ''
        description = ''
        
        for key, value in row_data.items():
            key_lower = key.lower()
            if 'назва органу' in key_lower or 'отримувач' in key_lower and 'банк' not in key_lower:
                recipient_name = value
            elif 'єдрпоу' in key_lower or 'код за' in key_lower or ('код' in key_lower and 'отримувач' in key_lower):
                recipient_code = value
            elif 'банк' in key_lower and 'номер' not in key_lower:
                bank_name = value
            elif 'iban' in key_lower or 'номер рахунку' in key_lower:
                iban = value
            elif 'категорії' in key_lower or 'примітк' in key_lower:
                description = value
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=None,
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=symbol,
            description=description
        )
        
        requisites.append(requisite)
    
    return requisites


def parse_tax_xls_simple(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Спрощений парсинг XLS файлу з фіксованою структурою
    
    Аналогічно parse_tax_xlsx_simple, але для старого формату Excel
    """
    workbook = xlrd.open_workbook(file_contents=file_content)
    sheet = workbook.sheet_by_index(0)
    
    requisites = []
    
    # Мапінг кодів класифікації до типів
    code_to_type = {
        '11010100': TaxRequisiteType.PDFO_EMPLOYEES.value,
        '11011000': TaxRequisiteType.MILITARY_EMPLOYEES.value,
        '11011700': TaxRequisiteType.MILITARY_FOP.value,
        '18050400': TaxRequisiteType.SINGLE_TAX_FOP.value,
    }
    
    # Мапінг регіонів до назв ГУК
    region_to_guk = {
        'Вінницька область': 'ГУК у Вінницькій області',
        'Волинська область': 'ГУК у Волинській області',
        'Дніпропетровська область': 'ГУК у Дніпропетровській області',
        'Донецька область': 'Донецьке ГУК',
        'Житомирська область': 'ГУК у Житомирській області',
        'Закарпатська область': 'ГУК у Закарпатській області',
        'Запорізька область': 'ГУК у Запорізькій області',
        'Івано-Франківська область': 'ГУК в Івано-Франківській області',
        'м. Київ': 'ГУК у м.Києві',
        'Київська область': 'ГУК у Київській області',
        'Кіровоградська область': 'ГУК у Кіровоградській області',
        'Луганська область': 'ГУК у Луганській області',
        'Львівська область': 'ГУК у Львівській області',
        'Миколаївська область': 'ГУК у Миколаївській області',
        'Одеська область': 'ГУК в Одеській області',
        'Полтавська область': 'ГУК у Полтавській області',
        'Рівненська область': 'ГУК у Рівненській області',
        'Сумська область': 'ГУК у Сумській області',
        'Тернопільська область': 'ГУК у Тернопільській області',
        'Харківська область': 'ГУК у Харківській області',
        'Херсонська область': 'ГУК у Херсонській області',
        'Хмельницька область': 'ГУК у Хмельницькій області',
        'Черкаська область': 'ГУК у Черкаській області',
        'Чернівецька область': 'ГУК у Чернівецькій області',
        'Чернігівська область': 'ГУК у Чернігівській області',
    }
    
    # Обробити всі рядки, починаючи з 2-го (індекс 1 в xlrd, бо нумерація з 0)
    for row_idx in range(1, sheet.nrows):
        row = sheet.row(row_idx)
        
        # Витягнути дані з фіксованих позицій
        district = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
        recipient_name_full = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
        recipient_code = str(row[2].value).strip().replace('.0', '') if len(row) > 2 and row[2].value else ''
        bank_name = str(row[3].value).strip() if len(row) > 3 and row[3].value else 'Казначейство України (ел. адм. подат.)'
        iban = str(row[4].value).strip() if len(row) > 4 and row[4].value else ''
        classification_code = str(row[5].value).strip().replace('.0', '') if len(row) > 5 and row[5].value else ''
        description = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''
        
        # Пропустити порожні рядки
        if not (district or recipient_name_full or iban):
            continue
        
        # Фільтр: тільки потрібні коди
        if classification_code not in code_to_type:
            continue
        
        # Визначити тип
        requisite_type = code_to_type[classification_code]
        
        # Витягнути recipient_name з повної назви (частина до першого /)
        if '/' in recipient_name_full:
            recipient_name = recipient_name_full.split('/')[0].strip()
        else:
            recipient_name = recipient_name_full
        
        # Нормалізувати recipient_name до повної назви ГУК
        recipient_name = region_to_guk.get(region, recipient_name)
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=district,
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=classification_code,
            description=description
        )
        
        requisites.append(requisite)
    
    print(f"📊 Статистика парсингу (Simple XLSX):")
    print(f"  Фактично створено записів: {len(requisites)}")
    
    return requisites


def parse_esv_xlsx_simple(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Спрощений парсинг XLSX файлу з реквізитами ЄСВ з фіксованою структурою
    
    Структура файлу:
    - Рядок 1: Заголовки
    - Рядок 2+: Дані
    - Колонка A: Банк отримувача
    - Колонка B: Назва органу ДПС (отримувач)
    - Колонка C: Код за ЄДРПОУ органу ДПС (код отримувача)
    - Колонка D: Номер рахунку (IBAN)
    - Колонка E: Символ звітності (201 або 204)
    - Колонка F: Категорії платників
    """
    workbook = load_workbook(BytesIO(file_content))
    sheet = workbook.active
    
    requisites = []
    
    # Обробити всі рядки, починаючи з 2-го (індекс 2 в openpyxl, бо нумерація з 1)
    for row_idx in range(2, sheet.max_row + 1):
        row = [cell.value for cell in sheet[row_idx]]
        
        # Витягнути дані з фіксованих позицій
        bank_name = str(row[0]).strip() if len(row) > 0 and row[0] else ''
        recipient_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        recipient_code = str(row[2]).strip().replace('.0', '') if len(row) > 2 and row[2] else ''
        iban = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        symbol = str(row[4]).strip().replace('.0', '') if len(row) > 4 and row[4] else ''
        description = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        
        # Пропустити порожні рядки
        if not (bank_name or recipient_name or iban):
            continue
        
        # Фільтр: тільки символи 201 та 204
        if symbol not in ['201', '204']:
            continue
        
        # Визначити тип
        if symbol == '201':
            requisite_type = TaxRequisiteType.ESV_FOP.value
        else:  # 204
            requisite_type = TaxRequisiteType.ESV_EMPLOYEES.value
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=None,  # ESV завжди для всієї області
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=symbol,
            description=description
        )
        
        requisites.append(requisite)
    
    print(f"📊 Статистика парсингу (Simple ESV XLSX):")
    print(f"  Фактично створено записів: {len(requisites)}")
    
    return requisites


def parse_esv_xls_simple(file_content: bytes, region: str) -> List[TaxRequisiteCreate]:
    """
    Спрощений парсинг XLS файлу з реквізитами ЄСВ з фіксованою структурою
    
    Структура файлу:
    - Рядок 1: Заголовки
    - Рядок 2+: Дані
    - Колонка A: Банк отримувача
    - Колонка B: Назва органу ДПС (отримувач)
    - Колонка C: Код за ЄДРПОУ органу ДПС (код отримувача)
    - Колонка D: Номер рахунку (IBAN)
    - Колонка E: Символ звітності (201 або 204)
    - Колонка F: Категорії платників
    """
    workbook = xlrd.open_workbook(file_contents=file_content)
    sheet = workbook.sheet_by_index(0)
    
    requisites = []
    
    # Обробити всі рядки, починаючи з 2-го (індекс 1, бо нумерація з 0)
    for row_idx in range(1, sheet.nrows):
        row = sheet.row(row_idx)
        
        # Витягнути дані з фіксованих позицій
        bank_name = str(row[0].value).strip() if len(row) > 0 else ''
        recipient_name = str(row[1].value).strip() if len(row) > 1 else ''
        recipient_code = str(row[2].value).strip().replace('.0', '') if len(row) > 2 else ''
        iban = str(row[3].value).strip() if len(row) > 3 else ''
        symbol = str(row[4].value).strip().replace('.0', '') if len(row) > 4 else ''
        description = str(row[5].value).strip() if len(row) > 5 else ''
        
        # Пропустити порожні рядки
        if not (bank_name or recipient_name or iban):
            continue
        
        # Фільтр: тільки символи 201 та 204
        if symbol not in ['201', '204']:
            continue
        
        # Визначити тип
        if symbol == '201':
            requisite_type = TaxRequisiteType.ESV_FOP.value
        else:  # 204
            requisite_type = TaxRequisiteType.ESV_EMPLOYEES.value
        
        # Пропустити якщо немає обов'язкових полів
        if not (recipient_name and recipient_code and iban):
            continue
        
        # Створити об'єкт
        requisite = TaxRequisiteCreate(
            region=region,
            type=requisite_type,
            district=None,  # ESV завжди для всієї області
            recipient_name=recipient_name,
            recipient_code=recipient_code,
            bank_name=bank_name,
            iban=iban,
            classification_code=symbol,
            description=description
        )
        
        requisites.append(requisite)
    
    print(f"📊 Статистика парсингу (Simple ESV XLS):")
    print(f"  Фактично створено записів: {len(requisites)}")
    
    return requisites
