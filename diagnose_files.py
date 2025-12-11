#!/usr/bin/env python3
"""
Скрипт для діагностики проблемних XLSX файлів
"""
import sys
from openpyxl import load_workbook
import xlrd

def diagnose_xlsx(file_path):
    """Діагностика XLSX файлу"""
    print(f"\n{'='*80}")
    print(f"ДІАГНОСТИКА: {file_path}")
    print('='*80)
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        print(f"Кількість рядків: {sheet.max_row}")
        print(f"Кількість колонок: {sheet.max_column}")
        
        # Показати перші 10 рядків
        print("\nПерші 10 рядків:")
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row = [str(cell.value) if cell.value is not None else '' for cell in sheet[row_idx]]
            print(f"Рядок {row_idx}: {row[:8]}")  # Перші 8 колонок
        
        # Пошук рядка з кодами класифікації
        print("\nПошук кодів класифікації (11010100, 11011000, 11011700, 18050400):")
        target_codes = ['11010100', '11011000', '11011700', '18050400']
        found_codes = []
        
        for row_idx in range(1, min(100, sheet.max_row + 1)):
            row = sheet[row_idx]
            for cell in row:
                cell_value = str(cell.value).strip() if cell.value is not None else ''
                # Видалити .0 якщо є
                cell_value = cell_value.replace('.0', '')
                if cell_value in target_codes and cell_value not in found_codes:
                    found_codes.append(cell_value)
                    print(f"  ✅ Знайдено код {cell_value} на рядку {row_idx}")
        
        if not found_codes:
            print("  ❌ НЕ ЗНАЙДЕНО жодного потрібного коду в перших 100 рядках")
            # Показати всі унікальні значення, що виглядають як коди (8 цифр)
            print("\nВсі значення що виглядають як коди (8 цифр):")
            unique_codes = set()
            for row_idx in range(1, min(100, sheet.max_row + 1)):
                row = sheet[row_idx]
                for cell in row:
                    cell_value = str(cell.value).strip() if cell.value is not None else ''
                    cell_value = cell_value.replace('.0', '')
                    if cell_value.isdigit() and len(cell_value) == 8:
                        unique_codes.add(cell_value)
            print(f"  {sorted(list(unique_codes))[:20]}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Помилка: {str(e)}")


def diagnose_xls(file_path):
    """Діагностика XLS файлу"""
    print(f"\n{'='*80}")
    print(f"ДІАГНОСТИКА XLS: {file_path}")
    print('='*80)
    
    try:
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        
        print(f"Кількість рядків: {sheet.nrows}")
        print(f"Кількість колонок: {sheet.ncols}")
        
        # Показати перші 10 рядків
        print("\nПерші 10 рядків:")
        for row_idx in range(min(10, sheet.nrows)):
            row = [str(cell.value).strip() for cell in sheet.row(row_idx)]
            print(f"Рядок {row_idx + 1}: {row[:8]}")
        
        # Пошук кодів класифікації
        print("\nПошук кодів класифікації:")
        target_codes = ['11010100', '11011000', '11011700', '18050400']
        found_codes = []
        
        for row_idx in range(min(100, sheet.nrows)):
            row = sheet.row(row_idx)
            for cell in row:
                cell_value = str(cell.value).strip().replace('.0', '')
                if cell_value in target_codes and cell_value not in found_codes:
                    found_codes.append(cell_value)
                    print(f"  ✅ Знайдено код {cell_value} на рядку {row_idx + 1}")
        
        if not found_codes:
            print("  ❌ НЕ ЗНАЙДЕНО жодного потрібного коду")
            
    except Exception as e:
        print(f"❌ Помилка: {str(e)}")


if __name__ == "__main__":
    files_to_check = [
        "/tmp/Івано-ФранківськаОбщ.xlsx",
        "/tmp/КиєвОбщ.xlsx",
        "/tmp/ДніпропетровськаОбщ.xlsx",
        "/tmp/ЖитомирськаОбщ.xlsx",
        "/tmp/ХерсонськаОбщ.xlsx",
    ]
    
    for file_path in files_to_check:
        if file_path.endswith('.xlsx'):
            # Перевірити формат
            with open(file_path, 'rb') as f:
                magic = f.read(8)
                if magic == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                    # Старий формат
                    diagnose_xls(file_path)
                else:
                    # Новий формат
                    diagnose_xlsx(file_path)

